import re
from datetime import datetime
from openpyxl import load_workbook
from pathlib import Path


class ExcelTaskReader:
    """
    Excelファイルからタスクデータを読み込むクラス
    """
    def __init__(self, config):
        self.config = config

    @staticmethod
    def extract_cell_data(sheet, row, date):
        """
        Excelのセルからデータを抽出する
        
        Parameters:
        -----------
        sheet : worksheet
            対象のワークシート
        row : int
            行番号
        date : datetime
            日付
            
        Returns:
        --------
        dict or None
            抽出されたデータ、抽出できない場合はNone
        """
        content = sheet[f'B{row}'].value
        time = sheet[f'C{row}'].value

        if not (content and time and time != '*'):
            return None

        try:
            # 数値に変換できる場合のみ処理する
            if isinstance(time, (int, float)):
                minutes = float(time)
            elif isinstance(time, str):
                try:
                    minutes = float(time)
                except ValueError:
                    return None
            else:
                return None

            content = content.split()[0] if content else content
            result = {
                'date': date,
                'content': content,
                'minutes': minutes
            }

            return result
        except (ValueError, TypeError) as e:
            print(f"時間の変換でエラー: {e}")
            return None

    def load_excel_task_data(self, wb, sheet_name, date):
        """
        ワークシートから各種タスクデータを読み込む
        
        Parameters:
        -----------
        wb : Workbook
            Excelワークブック
        sheet_name : str
            シート名
        date : datetime
            日付
            
        Returns:
        --------
        tuple
            (tasks, daily_tasks, communication_tasks)
        """
        sheet = wb[sheet_name]
        tasks = []
        daily_tasks = []
        communication_tasks = []

        # クラーク業務とクラーク業務以外のデータを抽出
        start_row = self.config.getint('Analysis', 'start_row')
        end_row = self.config.getint('Analysis', 'end_row')

        for row in range(start_row, end_row + 1):
            data = self.extract_cell_data(sheet, row, date)
            if data:
                tasks.append(data)

        # デイリータスクのデータを抽出
        daily_start_row = self.config.getint('Analysis', 'daily_task_start_row')
        daily_end_row = self.config.getint('Analysis', 'daily_task_end_row')

        for row in range(daily_start_row, daily_end_row + 1):
            data = self.extract_cell_data(sheet, row, date)
            if data:
                daily_tasks.append(data)

        # コミュニケーションのデータを抽出
        comm_start_row = self.config.getint('Analysis', 'communication_start_row')
        comm_end_row = self.config.getint('Analysis', 'communication_end_row')

        for row in range(comm_start_row, comm_end_row + 1):
            content = sheet[f'B{row}'].value
            time = sheet[f'C{row}'].value

            if content and time and time != '*':
                try:
                    # 名前を抽出 (括弧内の文字列を取得)
                    name_match = re.search(r'\((.*?)\)', content)
                    if name_match:
                        name = name_match.group(1)
                        content = re.sub(r'\(.*?\)', '', content).strip()
                        content = content.split()[0]
                        minutes = float(time)
                        communication_tasks.append({
                            'date': date,
                            'name': name,
                            'content': content,
                            'minutes': minutes
                        })
                except (ValueError, TypeError) as e:
                    print(f"コミュニケーションデータの時間の変換でエラー: {e}")

        return tasks, daily_tasks, communication_tasks

    def load_excel_sheet_all_items(self, wb, sheet_name, date):
        """
        ワークシートから全項目データを読み込む
        
        Parameters:
        -----------
        wb : Workbook
            Excelワークブック
        sheet_name : str
            シート名
        date : datetime
            日付
            
        Returns:
        --------
        list
            全項目のデータリスト
        """
        sheet = wb[sheet_name]
        all_items = []

        start_row = self.config.getint('Analysis', 'start_row')
        end_row = self.config.getint('Analysis', 'daily_task_end_row')

        for row in range(start_row, end_row + 1):
            data = self.extract_cell_data(sheet, row, date)
            if data:
                all_items.append(data)

        return all_items

    def read_workbook(self, file_path, start_date, end_date):
        """
        Excelワークブックから指定期間のデータを読み込む
        
        Parameters:
        -----------
        file_path : str
            Excelファイルのパス
        start_date : datetime
            開始日
        end_date : datetime
            終了日
            
        Returns:
        --------
        tuple
            (all_tasks, all_daily_tasks, all_communication_tasks, all_items, 
             actual_start_date, actual_end_date)
        """
        wb = load_workbook(filename=file_path)
        all_tasks = []
        all_daily_tasks = []
        all_communication_tasks = []
        all_items = []
        dates = []

        for sheet_name in wb.sheetnames:
            if sheet_name == 'シート一覧':
                continue

            sheet = wb[sheet_name]
            date_cell = sheet['A1'].value

            try:
                if isinstance(date_cell, datetime):
                    sheet_date = date_cell
                else:
                    sheet_date = datetime.strptime(str(date_cell), '%Y年%m月%d日')

                if start_date <= sheet_date <= end_date:
                    tasks, daily_tasks, comm_tasks = self.load_excel_task_data(wb, sheet_name, sheet_date)
                    sheet_items = self.load_excel_sheet_all_items(wb, sheet_name, sheet_date)

                    all_tasks.extend(tasks)
                    all_daily_tasks.extend(daily_tasks)
                    all_communication_tasks.extend(comm_tasks)
                    all_items.extend(sheet_items)
                    dates.append(sheet_date)

            except (ValueError, TypeError) as e:
                print(f"シート {sheet_name} の日付の解析でエラー: {e}")
                continue

        if not dates:
            raise ValueError("指定された期間内のデータがありません")

        actual_start_date = min(dates).strftime("%Y%m%d")
        actual_end_date = max(dates).strftime("%Y%m%d")

        return all_tasks, all_daily_tasks, all_communication_tasks, all_items, actual_start_date, actual_end_date
