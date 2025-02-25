import pytest
import tempfile
import configparser
from datetime import datetime
from openpyxl import Workbook
from service_excel_reader import ExcelTaskReader


@pytest.fixture
def mock_config():
    config = configparser.ConfigParser()
    config['Analysis'] = {
        'start_row': '5',
        'end_row': '15',
        'daily_task_start_row': '20',
        'daily_task_end_row': '25',
        'communication_start_row': '30',
        'communication_end_row': '35'
    }
    return config


@pytest.fixture
def mock_workbook():
    # テスト用のExcelファイルを作成
    wb = Workbook()
    
    # デフォルトシートを削除
    wb.remove(wb.active)
    
    # 3日分のシートを作成
    for day in range(1, 4):
        date_str = f"2024年1月{day}日"
        sheet = wb.create_sheet(title=f'シート{day}')
        
        # 日付を設定
        sheet['A1'] = date_str
        
        # クラーク業務データ
        if day == 1:
            sheet['B5'] = 'クラーク業務A'
            sheet['C5'] = 30
            sheet['B6'] = 'クラーク業務B'
            sheet['C6'] = 45
        elif day == 2:
            sheet['B5'] = 'クラーク業務A'
            sheet['C5'] = 25
            sheet['B6'] = '会議'
            sheet['C6'] = 60
        else:
            sheet['B5'] = '資料作成'
            sheet['C5'] = 90
        
        # デイリータスクデータ
        if day in [1, 2]:
            sheet['B20'] = '毎日タスクA'
            sheet['C20'] = 10
        if day in [1, 3]:
            sheet['B21'] = '毎日タスクB'
            sheet['C21'] = 15
        
        # コミュニケーションデータ
        if day == 1:
            sheet['B30'] = '打合せ(田中)'
            sheet['C30'] = 30
            sheet['B31'] = 'レビュー(佐藤)'
            sheet['C31'] = 45
        elif day == 2:
            sheet['B30'] = '打合せ(田中)'
            sheet['C30'] = 25
            sheet['B31'] = '相談(鈴木)'
            sheet['C31'] = 15
        else:
            sheet['B30'] = 'レビュー(佐藤)'
            sheet['C30'] = 30
    
    # 特殊なシートを追加
    sheet_list = wb.create_sheet(title='シート一覧')
    
    # 一時ファイルに保存
    with tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False) as tmp:
        wb.save(tmp.name)
        return tmp.name


class TestExcelTaskReader:
    def test_extract_cell_data(self, mock_config):
        reader = ExcelTaskReader(mock_config)
        
        # テスト用のワークブック作成
        wb = Workbook()
        sheet = wb.active
        
        # テストデータ設定
        sheet['B1'] = 'テストタスク'
        sheet['C1'] = 30
        
        # 無効なデータ
        sheet['B2'] = 'テストタスク'
        sheet['C2'] = '*'
        
        sheet['B3'] = None
        sheet['C3'] = 30
        
        sheet['B4'] = 'テストタスク'
        sheet['C4'] = 'abc'  # 数値に変換できない
        
        # テスト実行
        date = datetime(2024, 1, 1)
        result1 = reader.extract_cell_data(sheet, 1, date)
        result2 = reader.extract_cell_data(sheet, 2, date)
        result3 = reader.extract_cell_data(sheet, 3, date)
        result4 = reader.extract_cell_data(sheet, 4, date)
        
        # 検証
        assert result1 is not None
        assert result1['content'] == 'テストタスク'
        assert result1['minutes'] == 30
        assert result1['date'] == date
        
        # 無効なデータは None を返す
        assert result2 is None
        assert result3 is None
        assert result4 is None

    def test_load_excel_task_data(self, mock_config, mock_workbook):
        reader = ExcelTaskReader(mock_config)
        wb = self.__load_workbook(mock_workbook) # テスト用のワークブックを読み込む
        
        # 日付
        date = datetime(2024, 1, 1)
        
        # テスト実行
        tasks, daily_tasks, comm_tasks = reader.load_excel_task_data(wb, 'シート1', date)
        
        # 検証
        assert len(tasks) == 2
        assert tasks[0]['content'] == 'クラーク業務A'
        assert tasks[0]['minutes'] == 30
        
        assert len(daily_tasks) == 2
        assert daily_tasks[0]['content'] == '毎日タスクA'
        
        assert len(comm_tasks) == 2
        assert comm_tasks[0]['name'] == '田中'
        assert comm_tasks[0]['content'] == '打合せ'
        assert comm_tasks[0]['minutes'] == 30

    def test_load_excel_sheet_all_items(self, mock_config, mock_workbook):
        reader = ExcelTaskReader(mock_config)
        wb = self.__load_workbook(mock_workbook) # テスト用のワークブックを読み込む
        
        # 日付
        date = datetime(2024, 1, 1)
        
        # テスト実行
        all_items = reader.load_excel_sheet_all_items(wb, 'シート1', date)
        
        # 検証
        assert len(all_items) == 4  # クラーク業務2つ + デイリータスク2つ

    def test_read_workbook(self, mock_config, mock_workbook):
        reader = ExcelTaskReader(mock_config)
        
        # テスト実行
        start_date = datetime(2024, 1, 1)
        end_date = datetime(2024, 1, 3)
        tasks, daily_tasks, comm_tasks, all_items, actual_start_date, actual_end_date = reader.read_workbook(
            mock_workbook, start_date, end_date
        )
        
        # 検証
        assert len(tasks) == 5
        assert len(daily_tasks) == 4
        assert len(comm_tasks) == 5
        assert len(all_items) == 9
        
        assert actual_start_date == '20240101'
        assert actual_end_date == '20240103'
        
    def test_read_workbook_date_filter(self, mock_config, mock_workbook):
        reader = ExcelTaskReader(mock_config)
        
        # 日付範囲を制限してテスト
        start_date = datetime(2024, 1, 2)
        end_date = datetime(2024, 1, 2)
        tasks, daily_tasks, comm_tasks, all_items, actual_start_date, actual_end_date = reader.read_workbook(
            mock_workbook, start_date, end_date
        )
        
        # 検証 - 1日分のデータのみ
        assert len(tasks) == 2
        assert actual_start_date == '20240102'
        assert actual_end_date == '20240102'
        
    def __load_workbook(self, mock_workbook):
        """テスト用のヘルパーメソッド（ExcelTaskReaderの非公開メソッドを模倣）"""
        from openpyxl import load_workbook
        return load_workbook(filename=mock_workbook)
