import pytest
import os
import tempfile
from pathlib import Path
from datetime import datetime
import polars as pl
from openpyxl import load_workbook
from service_excel_writer import ExcelResultWriter


@pytest.fixture
def mock_analysis_results():
    # クラーク業務
    clerk_tasks = pl.DataFrame({
        'content': ['クラーク業務A', 'クラーク業務B'],
        'total_minutes': [55, 45],
        'total_hours': [0, 0],
        'frequency': [2, 1]
    })
    
    # クラーク業務以外
    non_clerk_tasks = pl.DataFrame({
        'content': ['会議', '資料作成'],
        'total_minutes': [60, 90],
        'total_hours': [1, 1],
        'frequency': [1, 1]
    })
    
    # デイリータスク
    daily_tasks = pl.DataFrame({
        'content': ['毎日タスクA', '毎日タスクB'],
        'total_minutes': [20, 30],
        'total_hours': [0, 0],
        'frequency': [2, 2]
    })
    
    # コミュニケーション（名前別）
    communication_by_name = pl.DataFrame({
        'name': ['田中', '佐藤', '鈴木'],
        'total_minutes': [55, 75, 15],
        'total_hours': [0, 1, 0],
        'frequency': [2, 2, 1]
    })
    
    # コミュニケーション（内容別）
    communication_by_content = pl.DataFrame({
        'content': ['打合せ', 'レビュー', '相談'],
        'name': ['田中', '佐藤', '鈴木'],
        'total_minutes': [55, 75, 15],
        'total_hours': [0, 1, 0],
        'frequency': [2, 2, 1]
    })
    
    # 全項目
    all_items_summary = pl.DataFrame({
        'content': ['クラーク業務A', '毎日タスクA', '会議', '資料作成'],
        'total_minutes': [30, 10, 60, 90],
        'total_hours': [0, 0, 1, 1],
        'frequency': [1, 1, 1, 1]
    })
    
    return (
        clerk_tasks,
        non_clerk_tasks,
        daily_tasks,
        communication_by_name,
        communication_by_content,
        all_items_summary
    )


@pytest.fixture
def mock_template():
    # 一時的なExcelテンプレートを作成
    with tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False) as tmp:
        wb = load_workbook(filename=pytest.importorskip('openpyxl').writer.excel.save_virtual_workbook(load_workbook()))
        
        # テンプレートに必要なシートを追加
        wb.create_sheet('クラーク業務')
        wb.create_sheet('クラーク以外業務')
        wb.create_sheet('デイリータスク')
        wb.create_sheet('コミュニケーション')
        wb.create_sheet('コミュニケーション内容')
        wb.create_sheet('全項目')
        
        # 各シートにヘッダーを追加
        for sheet_name in wb.sheetnames:
            if sheet_name != 'Sheet':
                sheet = wb[sheet_name]
                sheet.cell(row=1, column=1, value='content')
                sheet.cell(row=1, column=2, value='total_minutes')
                sheet.cell(row=1, column=3, value='total_hours')
                sheet.cell(row=1, column=4, value='frequency')
        
        wb.save(tmp.name)
        return tmp.name


@pytest.fixture
def mock_output_dir():
    # 一時的な出力ディレクトリを作成
    with tempfile.TemporaryDirectory() as tmp_dir:
        yield tmp_dir


class TestExcelResultWriter:
    def test_save_results(self, mock_analysis_results, mock_template, mock_output_dir, monkeypatch):
        # os.system呼び出しをモック化
        monkeypatch.setattr(os, 'system', lambda cmd: None)
        
        # テスト日付
        start_date = datetime(2024, 1, 1)
        end_date = datetime(2024, 1, 3)
        
        # テスト実行
        output_file = ExcelResultWriter.save_results(
            mock_analysis_results,
            mock_template,
            mock_output_dir,
            start_date,
            end_date
        )
        
        # 検証
        assert os.path.exists(output_file)
        
        # 生成されたExcelファイルを検証
        wb = load_workbook(filename=output_file)
        
        # 各シートの存在確認
        expected_sheets = ['クラーク業務', 'クラーク以外業務', 'デイリータスク', 
                          'コミュニケーション', 'コミュニケーション内容', '全項目']
        for sheet_name in expected_sheets:
            assert sheet_name in wb.sheetnames
        
        # クラーク業務シートのデータ確認
        clerk_sheet = wb['クラーク業務']
        assert clerk_sheet.cell(row=2, column=1).value == 'クラーク業務A'
        assert clerk_sheet.cell(row=2, column=2).value == 55
        
        # ファイル名の確認（日付フォーマット）
        assert '20240101_20240103' in output_file
