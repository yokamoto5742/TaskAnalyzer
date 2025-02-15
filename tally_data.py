import polars as pl
import pandas as pd
from openpyxl import load_workbook
from datetime import datetime
from pathlib import Path


def process_excel_sheet(wb, sheet_name, date):
    """
    指定されたシートから業務データを抽出する
    """
    sheet = wb[sheet_name]
    print(f"日付: {date}")

    # B3:C43の範囲からデータを取得
    tasks = []
    for row in range(3, 44):
        content = sheet[f'B{row}'].value
        time = sheet[f'C{row}'].value
        print(f"行 {row}: 内容={content}, 時間={time}")

        if content and time and time != '*':  # 両方のセルに値があり、時間が'*'でない場合のみ処理
            try:
                minutes = float(time)
                tasks.append({
                    'date': date,
                    'content': content,
                    'minutes': minutes
                })
                print(f"タスクを追加: {content} ({minutes}分)")
            except (ValueError, TypeError) as e:
                print(f"時間の変換でエラー: {e}")

    print(f"シート {sheet_name} から {len(tasks)} 個のタスクを抽出しました")
    return tasks


def analyze_workbook(file_path):
    """
    Excelファイルを分析し、業務内容と時間を集計する
    """
    print(f"ファイル {file_path} を読み込みます")
    wb = load_workbook(filename=file_path)
    print(f"シート一覧: {wb.sheetnames}")

    # 全シートのデータを収集
    all_tasks = []
    for sheet_name in wb.sheetnames:
        print(f"\nシート {sheet_name} を処理中...")
        sheet = wb[sheet_name]
        date_cell = sheet['A1'].value

        try:
            if isinstance(date_cell, datetime):
                date = date_cell
            else:
                date = datetime.strptime(str(date_cell), '%Y年%m月%d日（水）')

            if date >= datetime(2023, 4, 1) and date <= datetime(2024, 12, 31):
                tasks = process_excel_sheet(wb, sheet_name, date)
                all_tasks.extend(tasks)
            else:
                print(f"日付 {date} は対象期間外です")
        except (ValueError, TypeError) as e:
            print(f"シート {sheet_name} の日付の解析でエラー: {e}")
            continue

    print(f"\n合計 {len(all_tasks)} 個のタスクを収集しました")
    if not all_tasks:
        raise ValueError("タスクが1つも収集できませんでした。データの形式を確認してください。")

    # Polarsデータフレームを作成
    df = pl.DataFrame(all_tasks)
    print("データフレームのカラム:", df.columns)
    print("データフレームの最初の数行:")
    print(df.head())

    # 業務内容ごとの集計
    summary = df.group_by('content').agg([
        pl.col('minutes').sum().alias('total_minutes'),
        pl.len().alias('frequency')
    ]).sort('total_minutes', descending=True)

    # 月ごとの集計
    monthly = df.with_columns([
        pl.col('date').dt.strftime('%Y-%m').alias('month')
    ]).group_by('month').agg([
        pl.col('minutes').sum().alias('total_minutes')
    ]).sort('month')

    # 日付ごとの業務内容別集計
    daily = df.group_by(['date', 'content']).agg([
        pl.col('minutes').sum().alias('total_minutes')
    ]).sort(['date', 'total_minutes'], descending=[False, True])

    return summary, monthly, daily


def save_to_csv(summary, monthly, daily, output_dir):
    """
    分析結果をCSVファイルとして保存する
    """

    # 出力ディレクトリの作成
    output_dir = Path(output_dir)
    output_dir.mkdir(parents=True, exist_ok=True)

    # Polars DataFrameをpandas DataFrameに変換して保存
    summary.to_pandas().to_csv(output_dir / 'summary_by_task.csv', encoding='utf-8-sig', index=False)
    monthly.to_pandas().to_csv(output_dir / 'summary_by_month.csv', encoding='utf-8-sig', index=False)
    daily.to_pandas().to_csv(output_dir / 'summary_by_date_and_task.csv', encoding='utf-8-sig', index=False)

    print(f"CSVファイルを {output_dir} に保存しました")


def main():
    # 入力ファイルと出力ディレクトリのパスを設定
    file_path = r"C:\Shinseikai\playground\willdolist.xlsx"  # Excelファイルのパス
    output_dir = "analysis_output"  # 出力ディレクトリ

    # 分析の実行
    summary, monthly, daily = analyze_workbook(file_path)

    # 結果の保存
    save_to_csv(summary, monthly, daily, output_dir)
    print(f"分析結果を {output_dir} ディレクトリに保存しました。")


if __name__ == "__main__":
    main()
