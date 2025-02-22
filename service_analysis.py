import polars as pl
from datetime import datetime
from pathlib import Path
import os
import re
from openpyxl import load_workbook

from config_manager import load_config


class TaskAnalyzer:
    def __init__(self):
        self.config = load_config()
        self.paths_config = self.config['PATHS']

    @staticmethod
    def extract_cell_data(sheet, row, date):
        content = sheet[f'B{row}'].value
        time = sheet[f'C{row}'].value

        if not (content and time and time != '*'):
            return None

        try:
            # 数値に変換できる場合のみ処理する
            if isinstance(time, (int, float)):
                minutes = float(time)
            elif isinstance(time, str):
                try:
                    minutes = float(time)
                except ValueError:
                    return None
            else:
                return None

            content = content.split()[0] if content else content
            result = {
                'date': date,
                'content': content,
                'minutes': minutes
            }

            return result
        except (ValueError, TypeError) as e:
            print(f"時間の変換でエラー: {e}")
            return None

    def load_excel_task_data(self, wb, sheet_name, date):
        sheet = wb[sheet_name]
        tasks = []
        daily_tasks = []
        communication_tasks = []

        # クラーク業務とクラーク業務以外のデータを抽出
        start_row = self.config.getint('Analysis', 'start_row')
        end_row = self.config.getint('Analysis', 'end_row')

        for row in range(start_row, end_row + 1):
            data = self.extract_cell_data(sheet, row, date)
            if data:
                tasks.append(data)

        # デイリータスクのデータを抽出
        daily_start_row = self.config.getint('Analysis', 'daily_task_start_row')
        daily_end_row = self.config.getint('Analysis', 'daily_task_end_row')

        for row in range(daily_start_row, daily_end_row + 1):
            data = self.extract_cell_data(sheet, row, date)
            if data:
                daily_tasks.append(data)

        # コミュニケーションのデータを抽出
        comm_start_row = self.config.getint('Analysis', 'communication_start_row')
        comm_end_row = self.config.getint('Analysis', 'communication_end_row')

        for row in range(comm_start_row, comm_end_row + 1):
            content = sheet[f'B{row}'].value
            time = sheet[f'C{row}'].value

            if content and time and time != '*':
                try:
                    # 名前を抽出 (括弧内の文字列を取得)
                    name_match = re.search(r'\((.*?)\)', content)
                    if name_match:
                        name = name_match.group(1)
                        minutes = float(time)
                        communication_tasks.append({
                            'date': date,
                            'name': name,
                            'content': content,
                            'minutes': minutes
                        })
                except (ValueError, TypeError) as e:
                    print(f"コミュニケーションデータの時間の変換でエラー: {e}")

        return tasks, daily_tasks, communication_tasks

    def process_excel_sheet_all_items(self, wb, sheet_name, date):
        sheet = wb[sheet_name]
        all_items = []

        start_row = self.config.getint('Analysis', 'start_row')
        end_row = self.config.getint('Analysis', 'daily_task_end_row')

        for row in range(start_row, end_row + 1):
            data = self.extract_cell_data(sheet, row, date)
            if data:
                all_items.append(data)

        return all_items

    def analyze_workbook(self, file_path, start_date, end_date):
        wb = load_workbook(filename=file_path)
        all_tasks = []
        all_daily_tasks = []
        all_communication_tasks = []
        all_items = []
        dates = []

        for sheet_name in wb.sheetnames:
            if sheet_name == 'シート一覧':
                continue

            sheet = wb[sheet_name]
            date_cell = sheet['A1'].value

            try:
                if isinstance(date_cell, datetime):
                    sheet_date = date_cell
                else:
                    sheet_date = datetime.strptime(str(date_cell), '%Y年%m月%d日')

                if start_date <= sheet_date <= end_date:
                    tasks, daily_tasks, comm_tasks = self.load_excel_task_data(wb, sheet_name, sheet_date)
                    all_items = self.process_excel_sheet_all_items(wb, sheet_name, sheet_date)

                    all_tasks.extend(tasks)
                    all_daily_tasks.extend(daily_tasks)
                    all_communication_tasks.extend(comm_tasks)
                    all_items.extend(all_items)
                    dates.append(sheet_date)

            except (ValueError, TypeError) as e:
                print(f"シート {sheet_name} の日付の解析でエラー: {e}")
                continue

        if not dates:
            raise ValueError("有効なデータが見つかりませんでした。")

        df = pl.DataFrame(all_tasks)
        daily_df = pl.DataFrame(all_daily_tasks)
        comm_df = pl.DataFrame(all_communication_tasks)
        all_items_df = pl.DataFrame(all_items)

        start_date = min(dates).strftime("%Y%m%d")
        end_date = max(dates).strftime("%Y%m%d")

        return df, daily_df, comm_df, all_items_df, start_date, end_date

    @staticmethod
    def aggregate_dataframe(data_frame, group_by_col='content', filter_condition=None):
        if filter_condition is not None:
            data_frame = data_frame.filter(filter_condition)

        return (
            data_frame.group_by(group_by_col)
            .agg([
                pl.col('minutes').sum().alias('total_minutes'),
                (pl.col('minutes').sum() / 60).cast(pl.Int64).alias('total_hours'),
                pl.col('minutes').count().alias('frequency')
            ])
            .sort('total_minutes', descending=True)
        )

    def analyze_tasks(self, df, daily_df, comm_df, all_items_df, template_path, output_dir, start_date, end_date):
        clerk_tasks = self.aggregate_dataframe(
            df,
            filter_condition=pl.col('content').str.contains('クラーク業務')
        )
        non_clerk_tasks = self.aggregate_dataframe(
            df,
            filter_condition=~pl.col('content').str.contains('クラーク業務')
        )
        daily_tasks = self.aggregate_dataframe(daily_df)
        communication_summary = self.aggregate_dataframe(comm_df, group_by_col='name')
        all_items_summary = self.aggregate_dataframe(all_items_df)

        self.save_results(
            clerk_tasks,
            non_clerk_tasks,
            daily_tasks,
            communication_summary,
            all_items_summary,
            template_path,
            output_dir,
            start_date,
            end_date
        )

    @staticmethod
    def save_results(clerk_tasks, non_clerk_tasks, daily_tasks, communication_summary,
                     all_items_summary, template_path, output_dir, start_date, end_date):
        wb = load_workbook(filename=template_path)

        clerk_sheet = wb['クラーク業務']
        for i, row in enumerate(clerk_tasks.iter_rows(), start=2):
            for j, value in enumerate(row, start=1):
                clerk_sheet.cell(row=i, column=j, value=value)

        non_clerk_sheet = wb['クラーク以外業務']
        for i, row in enumerate(non_clerk_tasks.iter_rows(), start=2):
            for j, value in enumerate(row, start=1):
                non_clerk_sheet.cell(row=i, column=j, value=value)

        daily_sheet = wb['デイリータスク']
        for i, row in enumerate(daily_tasks.iter_rows(), start=2):
            for j, value in enumerate(row, start=1):
                daily_sheet.cell(row=i, column=j, value=value)

        comm_sheet = wb['コミュニケーション']
        for i, row in enumerate(communication_summary.iter_rows(), start=2):
            for j, value in enumerate(row, start=1):
                comm_sheet.cell(row=i, column=j, value=value)

        all_items_sheet = wb['全項目']
        for i, row in enumerate(all_items_summary.iter_rows(), start=2):
            for j, value in enumerate(row, start=1):
                all_items_sheet.cell(row=i, column=j, value=value)

        output_path = Path(output_dir)
        output_path.mkdir(parents=True, exist_ok=True)
        start_date_str = datetime.strftime(datetime.strptime(str(start_date), '%Y-%m-%d %H:%M:%S'), '%Y%m%d')
        end_date_str = datetime.strftime(datetime.strptime(str(end_date), '%Y-%m-%d %H:%M:%S'), '%Y%m%d')
        output_filename = f'WILLDOリストまとめ{start_date_str}_{end_date_str}.xlsx'
        output_file_path = output_path / output_filename

        wb.save(output_file_path)
        os.system(f'start excel "{output_file_path}"')

    def run_analysis(self, start_date_str, end_date_str):
        """分析処理の実行"""
        try:
            start_date = datetime.strptime(start_date_str, '%Y-%m-%d')
            end_date = datetime.strptime(end_date_str, '%Y-%m-%d')
            df, daily_df, comm_df, all_items_df, start_date_fmt, end_date_fmt = self.analyze_workbook(
                self.paths_config['input_file_path'],
                start_date,
                end_date
            )

            self.analyze_tasks(
                df,
                daily_df,
                comm_df,
                all_items_df,  # 全項目のDataFrameを追加
                self.paths_config['template_path'],
                self.paths_config['output_dir'],
                start_date,
                end_date
            )

            return True
        except Exception as e:
            print(f"エラーが発生しました: {e}")
            return False
