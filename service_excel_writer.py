import os
from pathlib import Path
from openpyxl import load_workbook


class ExcelResultWriter:
    @staticmethod
    def save_results(analysis_results, template_path, output_dir, start_date, end_date):
        clerk_tasks, non_clerk_tasks, daily_tasks, communication_by_name, \
        communication_by_content, all_items_summary = analysis_results

        wb = load_workbook(filename=template_path)

        clerk_sheet = wb['クラーク業務']
        for i, row in enumerate(clerk_tasks.iter_rows(), start=2):
            for j, value in enumerate(row, start=1):
                clerk_sheet.cell(row=i, column=j, value=value)

        non_clerk_sheet = wb['クラーク以外業務']
        for i, row in enumerate(non_clerk_tasks.iter_rows(), start=2):
            for j, value in enumerate(row, start=1):
                non_clerk_sheet.cell(row=i, column=j, value=value)

        daily_sheet = wb['デイリータスク']
        for i, row in enumerate(daily_tasks.iter_rows(), start=2):
            for j, value in enumerate(row, start=1):
                daily_sheet.cell(row=i, column=j, value=value)

        comm_sheet = wb['コミュニケーション']
        for i, row in enumerate(communication_by_name.iter_rows(), start=2):
            for j, value in enumerate(row, start=1):
                comm_sheet.cell(row=i, column=j, value=value)

        comm_content_sheet = wb['コミュニケーション内容']
        for i, row in enumerate(communication_by_content.iter_rows(), start=2):
            for j, value in enumerate(row, start=1):
                comm_content_sheet.cell(row=i, column=j, value=value)

        all_items_sheet = wb['全項目']
        for i, row in enumerate(all_items_summary.iter_rows(), start=2):
            for j, value in enumerate(row, start=1):
                all_items_sheet.cell(row=i, column=j, value=value)

        output_path = Path(output_dir)
        output_path.mkdir(parents=True, exist_ok=True)

        start_date_str = start_date.strftime('%Y%m%d')
        end_date_str = end_date.strftime('%Y%m%d')
        output_filename = f'WILLDOリストまとめ{start_date_str}_{end_date_str}.xlsx'
        output_file_path = output_path / output_filename

        wb.save(output_file_path)

        os.system(f'start excel "{output_file_path}"')
        
        return str(output_file_path)
