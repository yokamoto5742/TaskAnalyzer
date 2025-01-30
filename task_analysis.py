import polars as pl
import pandas as pd
from pathlib import Path
import datetime
import os
import math
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils.dataframe import dataframe_to_rows


def create_sheet_from_df(wb, sheet_name, df, start_row=1):
    """DataFrameをExcelシートに変換して書き込む"""
    ws = wb.create_sheet(title=sheet_name)

    # DataFrameの行をシートに書き込む
    for r_idx, row in enumerate(dataframe_to_rows(df, index=False), start_row):
        for c_idx, value in enumerate(row, 1):
            cell = ws.cell(row=r_idx, column=c_idx, value=value)

    # 列幅の自動調整
    for column in ws.columns:
        max_length = 0
        column_name = column[0].column_letter
        for cell in column:
            if cell.value:
                max_length = max(max_length, len(str(cell.value)))
        adjusted_width = (max_length + 2)
        ws.column_dimensions[column_name].width = adjusted_width


def analyze_tasks(file_path: str, output_dir: str) -> None:
    # CSVファイルの読み込み
    df = pl.read_csv(file_path, encoding='utf-8')

    # クラーク業務の集計
    clerk_tasks = (
        df.filter(pl.col('業務内容').str.contains('クラーク業務'))
        .group_by('業務内容')
        .agg([
            pl.col('時間').sum().alias('合計時間(分)'),
            (pl.col('時間').sum() / 60).cast(pl.Int64).alias('合計時間'),
            pl.col('時間').count().alias('タスク数')
        ])
        .sort('合計時間(分)', descending=True)
    )

    # クラーク以外の業務の集計
    non_clerk_tasks = (
        df.filter(~pl.col('業務内容').str.contains('クラーク業務'))
        .group_by('業務内容')
        .agg([
            pl.col('時間').sum().alias('合計時間(分)'),
            (pl.col('時間').sum() / 60).cast(pl.Int64).alias('合計時間'),
            pl.col('時間').count().alias('タスク数')
        ])
        .sort('合計時間(分)', descending=True)
    )

    # 全業務の集計
    all_tasks = (
        df.group_by('業務内容')
        .agg([
            pl.col('時間').sum().alias('合計時間(分)'),
            (pl.col('時間').sum() / 60).cast(pl.Int64).alias('合計時間'),
            pl.col('時間').count().alias('タスク数')
        ])
        .sort('合計時間(分)', descending=True)
    )

    # 各集計の合計を計算
    clerk_total = {
        '合計時間(分)': clerk_tasks['合計時間(分)'].sum(),
        '合計時間': math.floor(clerk_tasks['合計時間(分)'].sum() / 60),
        'タスク数': clerk_tasks['タスク数'].sum()
    }
    non_clerk_total = {
        '合計時間(分)': non_clerk_tasks['合計時間(分)'].sum(),
        '合計時間': math.floor(non_clerk_tasks['合計時間(分)'].sum() / 60),
        'タスク数': non_clerk_tasks['タスク数'].sum()
    }
    all_total = {
        '合計時間(分)': all_tasks['合計時間(分)'].sum(),
        '合計時間': math.floor(all_tasks['合計時間(分)'].sum() / 60),
        'タスク数': all_tasks['タスク数'].sum()
    }

    # 結果の表示
    print("\n=== クラーク業務の集計結果 ===")
    print(f"合計時間(分): {clerk_total['合計時間(分)']:.0f} 分")
    print(f"合計時間: {clerk_total['合計時間']:.0f} 時間")
    print(f"総タスク数: {clerk_total['タスク数']}件")

    print("\n=== クラーク以外の業務の集計結果 ===")
    print(f"合計時間(分): {non_clerk_total['合計時間(分)']:.0f} 分")
    print(f"合計時間: {non_clerk_total['合計時間']:.0f} 時間")
    print(f"総タスク数: {non_clerk_total['タスク数']}件")

    print("\n=== 全業務の集計結果 ===")
    print(f"合計時間(分): {all_total['合計時間(分)']:.0f} 分")
    print(f"合計時間: {all_total['合計時間']:.0f} 時間")
    print(f"総タスク数: {all_total['タスク数']}件")

    # 出力ディレクトリの作成
    output_path = Path(output_dir)
    output_path.mkdir(parents=True, exist_ok=True)

    # Excelファイルの作成
    wb = Workbook()
    wb.remove(wb.active)  # デフォルトシートの削除

    # 各DataFrameをpandasに変換
    clerk_df = clerk_tasks.to_pandas()
    non_clerk_df = non_clerk_tasks.to_pandas()
    all_df = all_tasks.to_pandas()

    # 合計行の追加
    clerk_df.loc['合計'] = ['総計', clerk_total['合計時間(分)'], clerk_total['合計時間'], clerk_total['タスク数']]
    non_clerk_df.loc['合計'] = ['総計', non_clerk_total['合計時間(分)'], non_clerk_total['合計時間'], non_clerk_total['タスク数']]
    all_df.loc['合計'] = ['総計', all_total['合計時間(分)'], all_total['合計時間'], all_total['タスク数']]

    # 各シートの作成
    create_sheet_from_df(wb, 'クラーク業務', clerk_df)
    create_sheet_from_df(wb, 'クラーク以外業務', non_clerk_df)
    create_sheet_from_df(wb, '全業務', all_df)

    # ファイルの保存
    current_time = datetime.datetime.now().strftime("%Y%m%d_%H%M%S")
    excel_file = output_path / f"tasks_summary_{current_time}.xlsx"
    wb.save(excel_file)
    print(f"\n集計結果を保存しました: {excel_file}")

    # Excelファイルを開く
    os.system(f"start excel {excel_file}")

if __name__ == "__main__":
    file_path = r"C:\Users\yokam\PycharmProjects\TaskAnalyzer\analysis_output\summary_by_task.csv"
    output_dir = "analysis_output"
    analyze_tasks(file_path, output_dir)
