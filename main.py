import polars as pl
import pandas as pd
from datetime import datetime
from pathlib import Path
import os
import math
from openpyxl import load_workbook, Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows


def process_excel_sheet(wb, sheet_name, date):
    sheet = wb[sheet_name]

    tasks = []
    for row in range(4, 24):
        content = sheet[f'B{row}'].value
        time = sheet[f'C{row}'].value

        if content and time and time != '*':
            try:
                content = content.split()[0] if content else content
                minutes = float(time)
                tasks.append({
                    'date': date,
                    'content': content,
                    'minutes': minutes
                })
            except (ValueError, TypeError) as e:
                print(f"時間の変換でエラー: {e}")

    return tasks


def analyze_workbook(file_path):
    wb = load_workbook(filename=file_path)

    all_tasks = []
    for sheet_name in wb.sheetnames:
        if sheet_name == 'シート一覧':
            continue

        sheet = wb[sheet_name]
        date_cell = sheet['A1'].value

        try:
            if isinstance(date_cell, datetime):
                date = date_cell
            else:
                date = datetime.strptime(str(date_cell), '%Y年%m月%d日')

            if datetime(2025, 1, 1) <= date <= datetime(2025, 12, 31):
                tasks = process_excel_sheet(wb, sheet_name, date)
                all_tasks.extend(tasks)

        except (ValueError, TypeError) as e:
            print(f"シート {sheet_name} の日付の解析でエラー: {e}")
            continue

    df = pl.DataFrame(all_tasks)

    summary = df.group_by('content').agg([
        pl.col('minutes').sum().alias('total_minutes'),
        pl.len().alias('frequency')
    ]).sort('total_minutes', descending=True)

    return summary


def save_to_csv(summary, output_dir):
    output_dir = Path(output_dir)
    output_dir.mkdir(parents=True, exist_ok=True)

    # 日本語のカラム名でCSVを保存
    renamed_summary = summary.select([
        pl.col('content').alias('業務内容'),
        pl.col('total_minutes').alias('時間'),
        pl.col('frequency').alias('タスク数')
    ])

    renamed_summary.to_pandas().to_csv(output_dir / 'summary_task.csv', encoding='utf-8-sig', index=False)


def apply_cell_style(cell, is_header=False, is_total=False):
    # 罫線のスタイル
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    cell.border = border

    # セルの配置
    cell.alignment = Alignment(horizontal='center', vertical='center')

    if is_header:
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color='D9D9D9', end_color='D9D9D9', fill_type='solid')
    elif is_total:
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color='FFE699', end_color='FFE699', fill_type='solid')


def create_sheet_from_df(wb, sheet_name, df, start_row=1):
    ws = wb.create_sheet(title=sheet_name)

    # ヘッダー行のスタイル設定
    headers = list(df.columns)
    for c_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=start_row, column=c_idx, value=header)
        apply_cell_style(cell, is_header=True)

    # データ行の書き込みとスタイル設定
    for r_idx, row in enumerate(df.values, start_row + 1):
        is_total = r_idx == len(df) + start_row
        for c_idx, value in enumerate(row, 1):
            cell = ws.cell(row=r_idx, column=c_idx, value=value)
            apply_cell_style(cell, is_total=is_total)

    # 列幅の自動調整
    for column in ws.columns:
        max_length = 0
        column_name = column[0].column_letter
        for cell in column:
            if cell.value:
                max_length = max(max_length, len(str(cell.value)))
        adjusted_width = (max_length + 2)
        ws.column_dimensions[column_name].width = adjusted_width


def analyze_tasks(csv_file_path: str, output_dir: str) -> None:
    # CSVファイルの読み込み
    df = pl.read_csv(csv_file_path, encoding='utf-8')

    # クラーク業務の集計
    clerk_tasks = (
        df.filter(pl.col('業務内容').str.contains('クラーク業務'))
        .group_by('業務内容')
        .agg([
            pl.col('時間').sum().alias('合計時間(分)'),
            (pl.col('時間').sum() / 60).cast(pl.Int64).alias('合計時間'),
            pl.col('時間').count().alias('タスク数')
        ])
        .sort('合計時間(分)', descending=True)
    )

    # クラーク以外の業務の集計
    non_clerk_tasks = (
        df.filter(~pl.col('業務内容').str.contains('クラーク業務'))
        .group_by('業務内容')
        .agg([
            pl.col('時間').sum().alias('合計時間(分)'),
            (pl.col('時間').sum() / 60).cast(pl.Int64).alias('合計時間'),
            pl.col('時間').count().alias('タスク数')
        ])
        .sort('合計時間(分)', descending=True)
    )

    # 全体の集計
    all_tasks = pl.concat([clerk_tasks, non_clerk_tasks])

    # 各集計の合計を計算
    clerk_total = {
        '合計時間(分)': clerk_tasks['合計時間(分)'].sum(),
        '合計時間': math.floor(clerk_tasks['合計時間(分)'].sum() / 60),
        'タスク数': clerk_tasks['タスク数'].sum()
    }

    non_clerk_total = {
        '合計時間(分)': non_clerk_tasks['合計時間(分)'].sum(),
        '合計時間': math.floor(non_clerk_tasks['合計時間(分)'].sum() / 60),
        'タスク数': non_clerk_tasks['タスク数'].sum()
    }

    all_total = {
        '合計時間(分)': all_tasks['合計時間(分)'].sum(),
        '合計時間': math.floor(all_tasks['合計時間(分)'].sum() / 60),
        'タスク数': all_tasks['タスク数'].sum()
    }

    output_path = Path(output_dir)
    output_path.mkdir(parents=True, exist_ok=True)

    wb = Workbook()
    wb.remove(wb.active)  # デフォルトシートの削除

    # DataFrameの変換
    clerk_df = clerk_tasks.to_pandas()
    non_clerk_df = non_clerk_tasks.to_pandas()
    all_df = all_tasks.to_pandas()

    # 合計行の追加
    clerk_df.loc['合計'] = ['総計', clerk_total['合計時間(分)'], clerk_total['合計時間'], clerk_total['タスク数']]
    non_clerk_df.loc['合計'] = ['総計', non_clerk_total['合計時間(分)'], non_clerk_total['合計時間'],
                                non_clerk_total['タスク数']]
    all_df.loc['合計'] = ['総計', all_total['合計時間(分)'], all_total['合計時間'], all_total['タスク数']]

    # 各シートの作成
    create_sheet_from_df(wb, 'クラーク業務', clerk_df)
    create_sheet_from_df(wb, 'クラーク以外業務', non_clerk_df)
    create_sheet_from_df(wb, '全体集計', all_df)

    # ファイルの保存
    current_time = datetime.now().strftime("%Y%m%d")
    excel_file = output_path / f"tasks_summary_{current_time}.xlsx"
    wb.save(excel_file)
    print(f"\n集計結果を保存しました: {excel_file}")

    # Excelファイルを開く
    os.system(f"start excel {excel_file}")


def main():
    file_path = r"C:\Shinseikai\TaskAnalyzer\WILLDOリスト.xlsx"
    output_dir = "analysis_output"
    summary = analyze_workbook(file_path)
    save_to_csv(summary, output_dir)
    csv_file_path = Path(output_dir) / 'summary_task.csv'
    analyze_tasks(str(csv_file_path), output_dir)


if __name__ == "__main__":
    main()
